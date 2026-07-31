"""
RadioEchoes Scraper - FAST Version
"""

import asyncio
import json
import logging
from pathlib import Path
from urllib.parse import urljoin

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright

# ================== CONFIG ==================
BASE_URL = "https://www.radioechoes.com"
START_URL = f"{BASE_URL}/?page=series_all"
HEADLESS = False          # False = Chrome dikhega
TIMEOUT = 45000
CONCURRENCY = 4           # parallel series (3-6 recommended)

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "logs"
STATE_DIR = BASE_DIR / "state"

OUTPUT_FILE = OUTPUT_DIR / "radioechoes.xlsx"
LOG_FILE = LOG_DIR / "scraper.log"
RESUME_FILE = STATE_DIR / "resume.json"

HEADERS = [
    "Series Name", "Episode Name", "Genre",
    "Original Broadcast Date", "Episode Length",
    "Download Link", "Play Link", "File Size"
]

# ================== SETUP ==================
def setup():
    OUTPUT_DIR.mkdir(exist_ok=True)
    LOG_DIR.mkdir(exist_ok=True)
    STATE_DIR.mkdir(exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )

    if not OUTPUT_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Radio Echoes"
        ws.append(HEADERS)
        wb.save(OUTPUT_FILE)

    if not RESUME_FILE.exists():
        with open(RESUME_FILE, "w", encoding="utf-8") as f:
            json.dump({"completed": [], "failed": []}, f, indent=2)


def load_resume():
    with open(RESUME_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_resume(data):
    with open(RESUME_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def append_to_excel(rows):
    if not rows:
        return
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])
    wb.save(OUTPUT_FILE)


# ================== SCRAPER ==================
async def get_series_links(page):
    await page.goto(START_URL, wait_until="domcontentloaded", timeout=TIMEOUT)
    hrefs = await page.eval_on_selector_all(
        "a[href*='page=series&']",
        "els => els.map(e => e.getAttribute('href'))"
    )
    links = {
        urljoin(BASE_URL, h)
        for h in hrefs
        if h and "series=" in h
    }
    return sorted(links)


async def scrape_series(context, url, logger):
    page = await context.new_page()
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT)

        data = await page.evaluate("""() => {
            let series = 'N/A';
            let genre = 'N/A';

            const allText = document.body.innerText;
            const sMatch = allText.match(/Series:\\s*(.+)/);
            if (sMatch) series = sMatch[1].split('\\n')[0].trim();

            const gMatch = allText.match(/Genre:\\s*(.+)/);
            if (gMatch) genre = gMatch[1].split('\\n')[0].trim();

            const rows = [];
            const playLinks = [...document.querySelectorAll("a[href*='mode=play']")];

            playLinks.forEach(a => {
                const parent = a.parentElement;
                if (!parent) return;

                const text = parent.innerText || '';
                const lines = text.split('\\n').map(l => l.trim()).filter(Boolean);

                let ep_name = 'N/A';
                let date = 'N/A';
                let length = 'N/A';
                let size = 'N/A';

                for (let i = 0; i < lines.length; i++) {
                    if (lines[i].includes('Original Broadcast Date')) {
                        date = lines[i].split(':').slice(1).join(':').trim();
                        if (i > 0) ep_name = lines[i - 1];
                    }
                    if (lines[i].includes('Length:')) {
                        length = lines[i].split('Length:')[1].trim();
                    }
                    if (lines[i].includes('File Size')) {
                        size = lines[i].split(':').slice(1).join(':').trim();
                    }
                }

                const dl = parent.querySelector("a[href*='mode=download']");
                const download = dl ? dl.href : 'N/A';
                const play = a.href || 'N/A';

                rows.push({
                    series, genre, ep_name, date, length, download, play, size
                });
            });

            return rows;
        }""")

        rows = []
        for item in data:
            rows.append({
                "Series Name": item["series"],
                "Episode Name": item["ep_name"],
                "Genre": item["genre"],
                "Original Broadcast Date": item["date"],
                "Episode Length": item["length"],
                "Download Link": item["download"],
                "Play Link": item["play"],
                "File Size": item["size"]
            })

        return rows

    except Exception as e:
        logger.error(f"Failed {url}: {e}")
        return None
    finally:
        await page.close()


# ================== MAIN ==================
async def main():
    setup()
    logger = logging.getLogger("RadioEchoes")
    resume = load_resume()
    done = set(resume["completed"])

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-sandbox"
            ]
        )
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 800}
        )

        page = await context.new_page()
        logger.info("Collecting series links...")
        links = await get_series_links(page)
        await page.close()
        logger.info(f"Total series: {len(links)}")

        pending = [u for u in links if u not in done]
        logger.info(f"Pending: {len(pending)}")

        sem = asyncio.Semaphore(CONCURRENCY)

        async def worker(url, idx, total):
            async with sem:
                logger.info(f"[{idx}/{total}] {url}")
                rows = await scrape_series(context, url, logger)

                if rows is None:
                    resume["failed"].append(url)
                    save_resume(resume)
                    return

                append_to_excel(rows)
                resume["completed"].append(url)
                save_resume(resume)
                logger.info(f"  → {len(rows)} episodes")

        tasks = [
            worker(url, i, len(pending))
            for i, url in enumerate(pending, 1)
        ]
        await asyncio.gather(*tasks)

        await browser.close()

    logger.info("All done!")


if __name__ == "__main__":
    asyncio.run(main())
